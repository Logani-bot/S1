"""
Trading Signal System - 20일 이평선 엔벨로프 기반 매매 시그널
- marketcap_universe.xlsx의 종목들을 분석 (S1 시스템)
- 3단계 분할 매수/매도 시그널 생성
- trading_signals.xlsx (Summary + History 탭) 생성
- 거래일(평일)에만 실행 (주말/공휴일 제외)
"""

import argparse
import logging
import time
import sys
from datetime import datetime, timedelta, date
from pathlib import Path
from typing import Dict, List, Optional, Tuple
import os
import pandas as pd
import requests
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill

# 거래일 체크 유틸리티 import
from trading_day_utils import is_trading_day, get_trading_day_info

# 텔레그램 알람
try:
    from telegram_notifier import send_daily_report, send_error_alert
    TELEGRAM_AVAILABLE = True
except ImportError:
    TELEGRAM_AVAILABLE = False
    logger_telegram = logging.getLogger(__name__)
    logger_telegram.warning("telegram_notifier 모듈을 찾을 수 없습니다. 텔레그램 알람이 비활성화됩니다.")

# ==================== 설정 ====================
logging.basicConfig(
    level=logging.INFO,
    format='[%(asctime)s] %(levelname)s: %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# API 설정
API_BASE_URL = "https://api.kiwoom.com"
API_TOKEN_URL = "https://api.kiwoom.com/oauth2/token"
API_CHART_ENDPOINT = "/api/dostk/chart"
API_CHART_ID = "ka10081"

# 기본 파일 경로
DEFAULT_UNIVERSE_FILE = "output/marketcap_universe.xlsx"
DEFAULT_SIGNAL_FILE = "output/trading_signals_s1.xlsx"
DEFAULT_ALERT_THRESHOLD = 10.0  # 알람 임계값 (%)

# 매수선 간격 (%)
BUY_LEVEL_GAP = 10.0  # 1차 → 2차 → 3차 각 10% 간격

# 매도선 간격 (%)
SELL_LEVELS = [3.0, 5.0, 7.0]  # +3%, +5%, +7%


# ==================== 호가 단위 계산 ====================
def get_tick_unit(price: float) -> int:
    """
    한국 주식시장 정확한 호가 단위 반환
    
    Args:
        price: 기준 가격
    
    Returns:
        호가 단위
    """
    if price < 2000:
        return 1
    elif price < 5000:
        return 5
    elif price < 20000:
        return 10
    elif price < 50000:
        return 50
    elif price < 200000:
        return 100
    elif price < 500000:
        return 500
    else:
        return 1000


def get_nearest_tick_price(price: float) -> float:
    """
    가장 가까운 정규 호가 가격 계산 (항상 한 호가 위)
    
    Args:
        price: 기준 가격
    
    Returns:
        가장 가까운 정규 호가 가격 (항상 한 호가 위)
    """
    tick_unit = get_tick_unit(price)
    
    # 현재 가격이 정확히 호가 단위에 맞는 경우도 한 호가 위로 계산
    if price % tick_unit == 0:
        return price + tick_unit
    
    # 현재 가격이 호가 단위 사이에 있는 경우 항상 윗 호가로 설정
    lower_tick = (price // tick_unit) * tick_unit
    upper_tick = lower_tick + tick_unit
    
    return upper_tick


def get_one_tick_up_price(price: float) -> float:
    """
    한 호가 위 가격 계산
    
    Args:
        price: 기준 가격
    
    Returns:
        한 호가 위 가격
    """
    nearest_tick = get_nearest_tick_price(price)
    tick_unit = get_tick_unit(nearest_tick)
    return nearest_tick + tick_unit


def predict_next_day_buy_price(S19_next: float) -> int:
    """
    다음날 매수 목표가 예측 (실시간 감시와 100% 동일한 로직)
    
    오늘(D일) 장마감 후 데이터로 다음날(D+1일) 실시간 감시 매수선과
    완전히 동일한 매수 목표가를 미리 계산
    
    Args:
        S19_next: 다음날 기준 19일 종가 합계 (S20 - Close_D_19)
    
    Returns:
        매수 목표가 p (정상 호가)
    """
    import math
    
    # 1. 이론적 접점 계산
    x_star = S19_next / 24.0
    
    # 2. 최초 후보 설정 (윗 호가로 올림)
    p = ceil_tick(x_star)
    
    # 3. 반복 검증
    while True:
        delta = get_tick_unit(p)                      # 호가 단위 재판정
        upper = (S19_next + 25.0 * delta) / 24.0      # 상한 계산
        
        if p < upper:                                 # 조건 충족 시 확정
            return int(p)
        else:
            p = p + delta                             # 조건 미충족 → 한 호가 위로 이동


def ceil_tick(price: float) -> float:
    """
    호가 단위에 맞춰 윗 호가로 올림
    
    Args:
        price: 기준 가격
    
    Returns:
        윗 호가로 올림된 가격
    """
    import math
    
    delta = get_tick_unit(price)
    return math.ceil(price / delta) * delta


def floor_tick(price: float) -> float:
    """
    호가 단위에 맞춰 아래 호가로 내림 (매도선용)
    
    Args:
        price: 기준 가격
    
    Returns:
        아래 호가로 내림된 가격
    """
    import math
    
    delta = get_tick_unit(price)
    return math.floor(price / delta) * delta


# ==================== API 함수 ====================
def get_api_token(appkey: str, secret: str, max_retry: int = 3) -> str:
    """API 토큰 획득"""
    headers = {"Content-Type": "application/json;charset=UTF-8"}
    body = {
        "grant_type": "client_credentials",
        "appkey": appkey,
        "secretkey": secret
    }
    
    for attempt in range(max_retry):
        try:
            response = requests.post(API_TOKEN_URL, headers=headers, json=body, timeout=20)
            response.raise_for_status()
            data = response.json()
            token = data.get("token") or data.get("access_token")
            
            if not token:
                raise ValueError("토큰을 찾을 수 없습니다")
            
            logger.info("✓ API 토큰 획득 완료")
            return token
            
        except Exception as e:
            if attempt == max_retry - 1:
                logger.error(f"토큰 획득 실패: {e}")
                raise
            logger.warning(f"토큰 획득 재시도 {attempt + 1}/{max_retry}")
            time.sleep(1)
    
    raise RuntimeError("토큰 획득 실패")


def fetch_chart_data(token: str, ticker: str, days: int = 60, max_retry: int = 5) -> pd.DataFrame:
    """차트 데이터 조회 (60일치) - KRX+NXT 통합 기준"""
    # ⭐ 키움 챗봇 권장: base_dt만 사용 (해당 날짜 이전 데이터 조회)
    base_date = datetime.now().strftime("%Y%m%d")
    
    # ⭐ KRX+NXT 통합 기준: 종목코드에 _AL 접미사 추가
    integrated_ticker = f"{ticker}_AL"
    
    headers = {
        "authorization": f"Bearer {token}",
        "Content-Type": "application/json;charset=UTF-8",
        "api-id": API_CHART_ID,
        "cont-yn": "N",
        "next-key": ""
    }
    
    body = {
        "stk_cd": integrated_ticker,  # 통합 종목코드 사용
        "base_dt": base_date,  # 오늘 날짜만 (end_dt 제거!)
        "upd_stkpc_tp": "1",  # 수정주가
        "stex_tp": "3"  # 통합 (KRX+NXT)
    }
    
    url = API_BASE_URL + API_CHART_ENDPOINT
    
    for attempt in range(max_retry):
        try:
            response = requests.post(url, headers=headers, json=body, timeout=20)
            response.raise_for_status()
            result = response.json()
            
            # 데이터 추출
            records = None
            for value in result.values():
                if isinstance(value, list) and len(value) > 0:
                    records = value
                    break
            
            if not records:
                logger.warning(f"  ⚠️ 차트 데이터 없음: {ticker}")
                return pd.DataFrame()
            
            # 데이터 파싱
            rows = []
            for rec in records:
                # 날짜
                dt_str = rec.get("dt") or rec.get("stck_bsop_date") or rec.get("bas_dd")
                if not dt_str:
                    continue
                
                # 가격 (종가 우선순위: END_PRC > stck_clpr > close > cur_prc)
                close_price = safe_float(rec.get("END_PRC") or rec.get("stck_clpr") or rec.get("close") or rec.get("cur_prc"))
                high_price = safe_float(rec.get("HIGH_PRC") or rec.get("stck_hgpr") or rec.get("high_pric") or rec.get("high"))
                low_price = safe_float(rec.get("LOW_PRC") or rec.get("stck_lwpr") or rec.get("low_pric") or rec.get("low"))
                
                if close_price and high_price and low_price:
                    rows.append({
                        "날짜": dt_str,
                        "종가": close_price,
                        "고가": high_price,
                        "저가": low_price
                    })
            
            if not rows:
                logger.warning(f"  ⚠️ 유효한 차트 데이터 없음: {ticker}")
                return pd.DataFrame()
            
            df = pd.DataFrame(rows)
            df["날짜"] = pd.to_datetime(df["날짜"])
            
            # ⭐ 날짜 내림차순 정렬 (최신이 먼저)
            df = df.sort_values("날짜", ascending=False).reset_index(drop=True)
            
            # 최근 days일만 (영업일 기준)
            df = df.head(days)
            
            # 다시 오름차순으로 정렬 (시계열 분석용)
            df = df.sort_values("날짜").reset_index(drop=True)
            
            return df
            
        except Exception as e:
            if attempt == max_retry - 1:
                logger.error(f"  ❌ 차트 조회 실패: {ticker} - {e}")
                return pd.DataFrame()
            time.sleep(0.5 * (2 ** attempt))
    
    return pd.DataFrame()


def safe_float(value) -> Optional[float]:
    """안전하게 float 변환"""
    if value is None or value == "":
        return None
    try:
        return float(str(value).replace(",", ""))
    except (ValueError, TypeError):
        return None


# ==================== 기술 지표 계산 ====================
def calculate_ma(df: pd.DataFrame, period: int = 20) -> float:
    """이동평균선 계산"""
    if len(df) < period:
        return None
    return df["종가"].tail(period).mean()


def calculate_envelope_support(ma: float, envelope_pct: float = -20.0) -> float:
    """엔벨로프 지지선 계산"""
    if ma is None:
        return None
    return ma * (1 + envelope_pct / 100)


def calculate_buy_line_1(S19_next: float) -> float:
    """
    1차 매수선: 다음날 실시간 감시와 100% 동일한 매수 목표가
    
    Args:
        S19_next: 다음날 기준 19일 종가 합계 (S20 - Close_D_19)
    
    Returns:
        1차 매수선 가격
    """
    if S19_next is None or S19_next <= 0:
        return None
    return predict_next_day_buy_price(S19_next)


def calculate_buy_line_2(buy1: float) -> float:
    """2차 매수선: 1차 매수선에서 10% 하락 후 가장 가까운 윗 호가"""
    if buy1 is None:
        return None
    base_price = buy1 * (1 - BUY_LEVEL_GAP / 100)
    return get_nearest_tick_price(base_price)


def calculate_buy_line_3(buy2: float) -> float:
    """3차 매수선: 2차 매수선에서 10% 하락 후 가장 가까운 윗 호가"""
    if buy2 is None:
        return None
    base_price = buy2 * (1 - BUY_LEVEL_GAP / 100)
    return get_nearest_tick_price(base_price)


def calculate_sell_line_1(avg_buy_price: float) -> float:
    """1차 매도선: 평균 매수가에서 3% 상승 후 아래 호가"""
    if avg_buy_price is None:
        return None
    target_price = avg_buy_price * (1 + SELL_LEVEL_1_GAP / 100)
    return floor_tick(target_price)


def calculate_sell_line_2(avg_buy_price: float) -> float:
    """2차 매도선: 평균 매수가에서 5% 상승 후 아래 호가"""
    if avg_buy_price is None:
        return None
    target_price = avg_buy_price * (1 + SELL_LEVEL_2_GAP / 100)
    return floor_tick(target_price)


def calculate_sell_line_3(avg_buy_price: float) -> float:
    """3차 매도선: 평균 매수가에서 7% 상승 후 아래 호가"""
    if avg_buy_price is None:
        return None
    target_price = avg_buy_price * (1 + SELL_LEVEL_3_GAP / 100)
    return floor_tick(target_price)


def calculate_distance_pct(current: float, target: float) -> float:
    """이격도 계산 (%) - 부동소수점 오차 보정"""
    if current is None or target is None or target == 0:
        return None
    
    # 부동소수점 오차 보정 (매우 작은 값은 0으로 처리)
    distance_pct = ((current - target) / target) * 100
    
    # 절댓값이 1e-10보다 작으면 0으로 처리 (극소값 제거)
    if abs(distance_pct) < 1e-10:
        return 0.0
    
    return distance_pct


# ==================== 매수/매도 로직 ====================
class BuyStatus:
    NONE = "NONE"
    BOUGHT_1 = "BOUGHT_1"
    BOUGHT_2 = "BOUGHT_2"
    BOUGHT_3 = "BOUGHT_3"
    SOLD = "SOLD"


class AlertStatus:
    WATCHING = "WATCHING"
    READY_BUY1 = "READY_BUY1"
    READY_BUY2 = "READY_BUY2"
    READY_BUY3 = "READY_BUY3"
    WAITING = "WAITING"
    READY_SELL1 = "READY_SELL1"
    READY_SELL2 = "READY_SELL2"
    READY_SELL3 = "READY_SELL3"
    COMPLETED = "COMPLETED"


def check_buy_signal(low: float, buy_line: float) -> bool:
    """매수 시그널 체크 (당일 저가가 매수선 이하)"""
    if low is None or buy_line is None:
        return False
    return low <= buy_line


def check_sell_retouch(high: float, close: float, sell_line: float, max_high: float, avg_buy_price: float) -> bool:
    """매도 재터치 시그널 체크 (매수가 기준)"""
    # 과거에 매도선 이상 도달했었고, 현재가가 평균매수가 근처로 재터치
    if max_high is None or high is None or close is None or sell_line is None or avg_buy_price is None:
        return False

    # 과거 최고가가 매도선 이상이었고
    if max_high >= sell_line:
        # 현재 종가가 평균매수가 근처 재터치 (±1% 허용)
        if abs(close - avg_buy_price) / avg_buy_price < 0.01:
            return True

    return False


# ==================== 종목 분석 ====================
def analyze_stock(token: str, ticker: str, name: str, df_summary: pd.DataFrame, alert_threshold: float) -> Optional[Dict]:
    """종목 분석 및 시그널 생성"""
    
    # 차트 데이터 조회
    df_chart = fetch_chart_data(token, ticker, days=60)
    
    if df_chart.empty:
        logger.warning(f"  ⚠️ 차트 데이터 없음 - 스킵")
        return None
    
    if len(df_chart) < 20:
        logger.warning(f"  ⚠️ 데이터 부족 ({len(df_chart)}일) - 20일 이상 필요")
        return None
    
    # 최신 데이터
    latest = df_chart.iloc[-1]
    
    close = latest["종가"]
    low = latest["저가"]
    high = latest["고가"]
    date_str = latest["날짜"].strftime("%Y-%m-%d")
    
    # 데이터 오래된 정도 확인
    now = datetime.now()
    latest_date = latest["날짜"].date()
    days_old = (now.date() - latest_date).days
    
    if days_old > 5:
        logger.warning(f"  ⚠️ 데이터가 오래됨: {date_str} ({days_old}일 전)")
    elif days_old > 0:
        logger.info(f"  📅 데이터 날짜: {date_str} ({days_old}일 전)")
    
    # ==================== 당일(D일) 기준 계산 ====================
    # 당일 20일선 계산
    ma20_today = calculate_ma(df_chart, 20)
    if ma20_today is None:
        logger.warning(f"  ⚠️ 20일선 계산 실패")
        return None

    # 당일 기준 19일 종가 합계 계산
    S20_today = ma20_today * 20  # 당일 20일 종가 합계
    Close_D_20 = df_chart.iloc[-20]["종가"]  # 20일 전 종가 (당일 기준)
    S19_today = S20_today - Close_D_20  # 당일 기준 19일 종가 합계

    # 당일 기준 매수선 계산 (체결 판단용)
    buy1_today = calculate_buy_line_1(S19_today)
    buy2_today = calculate_buy_line_2(buy1_today)
    buy3_today = calculate_buy_line_3(buy2_today)

    # ==================== 익일(D+1일) 기준 계산 ====================
    # 익일 기준 19일 종가 합계 계산
    # 익일 20일선 = (D-18 ~ D) 19일 + 익일종가
    # 따라서 S19_next = S20_today - D-19일 종가
    Close_D_19 = df_chart.iloc[-20]["종가"]  # D-19일 종가 (iloc[-20]이 20일 전 = D-19일)
    S19_next = S20_today - Close_D_19  # 익일 기준: 19일 전 종가 제외

    # 익일 기준 매수선 계산 (알람용)
    buy1_next = calculate_buy_line_1(S19_next)
    buy2_next = calculate_buy_line_2(buy1_next)
    buy3_next = calculate_buy_line_3(buy2_next)

    # 이격도 계산 (익일 기준)
    dist_buy1 = calculate_distance_pct(close, buy1_next)
    dist_buy2 = calculate_distance_pct(close, buy2_next)
    dist_buy3 = calculate_distance_pct(close, buy3_next)

    # 엔벨로프 지지선 (-20%) - 당일/익일 기준 각각 계산
    envelope_today = calculate_envelope_support(ma20_today, -20.0)

    # 익일 기준 20일선 계산 (D+1일에 D-19일 데이터가 빠지고 익일 종가가 추가됨)
    # 익일 S20 = S19_next + 익일종가
    # 익일 종가를 모르므로 당일 종가로 근사
    S20_next_approx = S19_next + close
    ma20_next_approx = S20_next_approx / 20
    envelope_next = calculate_envelope_support(ma20_next_approx, -20.0)

    logger.info(f"  [{date_str}] 종가: {close:,.0f}원, 20일선: {ma20_today:,.0f}원")
    logger.info(f"  📊 [당일 기준] S19_today: {S19_today:,.0f}")
    logger.info(f"  🎯 [당일 기준 매수선] 1차: {buy1_today:,.0f}, 2차: {buy2_today:,.0f}, 3차: {buy3_today:,.0f}")
    logger.info(f"  📊 [당일 기준 엔벨로프] -20%: {envelope_today:,.0f}원")
    logger.info(f"  📊 [익일 기준] S19_next: {S19_next:,.0f}")
    logger.info(f"  🎯 [익일 기준 매수선] 1차: {buy1_next:,.0f}, 2차: {buy2_next:,.0f}, 3차: {buy3_next:,.0f}")
    logger.info(f"  📊 [익일 기준 엔벨로프] -20%: {envelope_next:,.0f}원 (근사값)")
    
    # 기존 데이터 확인
    existing = df_summary[df_summary["티커"] == ticker] if not df_summary.empty and "티커" in df_summary.columns else pd.DataFrame()
    
    if existing.empty:
        # 신규 종목
        buy_status = BuyStatus.NONE
        avg_price = None
        total_qty = 0
        total_amount = 0
        buy1_date = None
        buy1_price = None
        buy1_qty = None
        buy2_date = None
        buy2_price = None
        buy2_qty = None
        buy3_date = None
        buy3_price = None
        buy3_qty = None
        max_high_line = None
    else:
        # 기존 종목
        row = existing.iloc[0]
        buy_status = row.get("매수상태", BuyStatus.NONE)
        avg_price = row.get("평균매수가")
        total_qty = row.get("총보유수량", 0)
        total_amount = row.get("총투자금액", 0)
        buy1_date = row.get("1차매수일")
        buy1_price = row.get("1차매수가")
        buy1_qty = row.get("1차매수량")
        buy2_date = row.get("2차매수일")
        buy2_price = row.get("2차매수가")
        buy2_qty = row.get("2차매수량")
        buy3_date = row.get("3차매수일")
        buy3_price = row.get("3차매수가")
        buy3_qty = row.get("3차매수량")
        max_high_line = row.get("최고도달선")
    
    # ==================== 매수 시그널 체크 (⭐ 당일 기준) ====================
    if buy_status == BuyStatus.NONE and check_buy_signal(low, buy1_today):
        buy_status = BuyStatus.BOUGHT_1
        buy1_date = date_str
        buy1_price = buy1_today  # ⭐ 당일 기준 매수가
        buy1_qty = 100  # 예시: 100주
        total_qty = 100
        total_amount = buy1_today * 100
        avg_price = buy1_today
        logger.info(f"  🔴 1차 매수 체결! {buy1_today:,.0f}원 x 100주 (당일 기준)")

    elif buy_status == BuyStatus.BOUGHT_1 and check_buy_signal(low, buy2_today):
        buy_status = BuyStatus.BOUGHT_2
        buy2_date = date_str
        buy2_price = buy2_today  # ⭐ 당일 기준 매수가
        buy2_qty = 100
        total_qty += 100
        total_amount += buy2_today * 100
        avg_price = total_amount / total_qty
        logger.info(f"  🔴🔴 2차 매수 체결! {buy2_today:,.0f}원 x 100주 (당일 기준)")

    elif buy_status == BuyStatus.BOUGHT_2 and check_buy_signal(low, buy3_today):
        buy_status = BuyStatus.BOUGHT_3
        buy3_date = date_str
        buy3_price = buy3_today  # ⭐ 당일 기준 매수가
        buy3_qty = 100
        total_qty += 100
        total_amount += buy3_today * 100
        avg_price = total_amount / total_qty
        logger.info(f"  🔴🔴🔴 3차 매수 체결! {buy3_today:,.0f}원 x 100주 (당일 기준)")
    
    # 매도선 계산 (매수 후에만)
    sell1 = None
    sell2 = None
    sell3 = None
    dist_sell1 = None
    dist_sell2 = None
    dist_sell3 = None
    
    if buy_status in [BuyStatus.BOUGHT_1, BuyStatus.BOUGHT_2, BuyStatus.BOUGHT_3] and avg_price:
        # avg_price가 문자열인 경우 처리
        if isinstance(avg_price, str):
            try:
                avg_price = float(avg_price.replace(",", ""))
            except (ValueError, TypeError):
                avg_price = 0
        
        if avg_price > 0:
            sell1 = avg_price * (1 + SELL_LEVELS[0] / 100)  # +3%
            sell2 = avg_price * (1 + SELL_LEVELS[1] / 100)  # +5%
            sell3 = avg_price * (1 + SELL_LEVELS[2] / 100)  # +7%
        else:
            sell1 = sell2 = sell3 = 0
        
        dist_sell1 = calculate_distance_pct(close, sell1)
        dist_sell2 = calculate_distance_pct(close, sell2)
        dist_sell3 = calculate_distance_pct(close, sell3)
        
        # 최고도달선 업데이트
        if max_high_line is None:
            max_high_line = high
        else:
            max_high_line = max(max_high_line, high)
        
        # 매도 시그널 체크
        # +7% 즉시 매도
        if high >= sell3:
            buy_status = BuyStatus.SOLD
            logger.info(f"  💰💰💰 +7% 도달! 전량 매도!")

        # +5% 도달 후 평균매수가 재터치
        elif check_sell_retouch(high, close, sell2, max_high_line, avg_price):
            buy_status = BuyStatus.SOLD
            logger.info(f"  💰💰 +5% 도달 후 매수가 재터치! 전량 매도!")

        # +3% 도달 후 평균매수가 재터치
        elif check_sell_retouch(high, close, sell1, max_high_line, avg_price):
            buy_status = BuyStatus.SOLD
            logger.info(f"  💰 +3% 도달 후 매수가 재터치! 전량 매도!")
    
    # 알람 상태 결정 (⭐ 익일 기준 매수선 사용)
    alert_status, alert_msg = determine_alert_status(
        buy_status, close, buy1_next, buy2_next, buy3_next, sell1, sell2, sell3,
        dist_buy1, dist_buy2, dist_buy3, dist_sell1, dist_sell2, dist_sell3,
        alert_threshold
    )

    # 결과 반환 (⭐ 당일/익일 기준 모두 Excel에 기록)
    result = {
        "티커": ticker,
        "종목명": name,
        "매수상태": buy_status,
        "알람상태": alert_status,
        "상태메시지": alert_msg,
        "종가": close,
        "저가": low,
        "고가": high,
        "20일선(당일)": ma20_today,
        "20일선(익일)": ma20_next_approx,
        "-20%엔벨로프(당일)": envelope_today,
        "-20%엔벨로프(익일)": envelope_next,
        "1차매수선(익일)": buy1_next,  # ⭐ 익일 기준 (알람용)
        "1차매수선이격도(%)": dist_buy1,
        "1차매수일": buy1_date,
        "1차매수가(당일)": buy1_price,  # ⭐ 당일 실제 체결가
        "1차매수량": buy1_qty,
        "2차매수선(익일)": buy2_next,  # ⭐ 익일 기준 (알람용)
        "2차매수선이격도(%)": dist_buy2,
        "2차매수일": buy2_date,
        "2차매수가(당일)": buy2_price,  # ⭐ 당일 실제 체결가
        "2차매수량": buy2_qty,
        "3차매수선(익일)": buy3_next,  # ⭐ 익일 기준 (알람용)
        "3차매수선이격도(%)": dist_buy3,
        "3차매수일": buy3_date,
        "3차매수가(당일)": buy3_price,  # ⭐ 당일 실제 체결가
        "3차매수량": buy3_qty,
        "평균매수가": avg_price,
        "총투자금액": total_amount,
        "총보유수량": total_qty,
        "1차매도선(+3%)": sell1,
        "1차매도선이격도(%)": dist_sell1,
        "2차매도선(+5%)": sell2,
        "2차매도선이격도(%)": dist_sell2,
        "3차매도선(+7%)": sell3,
        "3차매도선이격도(%)": dist_sell3,
        "최고도달선": max_high_line,
    }
    
    
    return result


def determine_alert_status(buy_status: str, close: float,
                           buy1: float, buy2: float, buy3: float,
                           sell1: float, sell2: float, sell3: float,
                           dist_buy1: float, dist_buy2: float, dist_buy3: float,
                           dist_sell1: float, dist_sell2: float, dist_sell3: float,
                           threshold: float) -> Tuple[str, str]:
    """알람 상태 및 메시지 결정"""
    
    # 매도선 접근 기준 (3%)
    SELL_ALERT_THRESHOLD = 3.0
    
    if buy_status == BuyStatus.SOLD:
        return AlertStatus.COMPLETED, "매도 완료"
    
    # 매수 전
    if buy_status == BuyStatus.NONE:
        if dist_buy1 is not None and 0 < dist_buy1 <= threshold:
            return AlertStatus.READY_BUY1, f"1차 매수선까지 {dist_buy1:.1f}% (접근 중!)"
        else:
            return AlertStatus.WATCHING, f"1차 매수선까지 {dist_buy1:.1f}%"
    
    # 1차 매수 후
    elif buy_status == BuyStatus.BOUGHT_1:
        # 매도선 체크 (3% 기준)
        if dist_sell1 is not None and abs(dist_sell1) <= SELL_ALERT_THRESHOLD:
            return AlertStatus.READY_SELL1, f"+3% 매도선까지 {abs(dist_sell1):.1f}%"
        # 2차 매수선 체크
        elif dist_buy2 is not None and 0 < dist_buy2 <= threshold:
            return AlertStatus.READY_BUY2, f"2차 매수선까지 {dist_buy2:.1f}%"
        else:
            return AlertStatus.WAITING, f"대기 중 (2차선까지 {dist_buy2:.1f}%)"
    
    # 2차 매수 후
    elif buy_status == BuyStatus.BOUGHT_2:
        # 매도선 체크 (3% 기준)
        if dist_sell2 is not None and abs(dist_sell2) <= SELL_ALERT_THRESHOLD:
            return AlertStatus.READY_SELL2, f"+5% 매도선까지 {abs(dist_sell2):.1f}%"
        elif dist_sell1 is not None and abs(dist_sell1) <= SELL_ALERT_THRESHOLD:
            return AlertStatus.READY_SELL1, f"+3% 매도선까지 {abs(dist_sell1):.1f}%"
        # 3차 매수선 체크
        elif dist_buy3 is not None and 0 < dist_buy3 <= threshold:
            return AlertStatus.READY_BUY3, f"3차 매수선까지 {dist_buy3:.1f}%"
        else:
            return AlertStatus.WAITING, f"대기 중 (3차선까지 {dist_buy3:.1f}%)"
    
    # 3차 매수 후
    elif buy_status == BuyStatus.BOUGHT_3:
        # 매도선 체크 (3% 기준)
        if dist_sell3 is not None and abs(dist_sell3) <= SELL_ALERT_THRESHOLD:
            return AlertStatus.READY_SELL3, f"+7% 매도선까지 {abs(dist_sell3):.1f}%"
        elif dist_sell2 is not None and abs(dist_sell2) <= SELL_ALERT_THRESHOLD:
            return AlertStatus.READY_SELL2, f"+5% 매도선까지 {abs(dist_sell2):.1f}%"
        elif dist_sell1 is not None and abs(dist_sell1) <= SELL_ALERT_THRESHOLD:
            return AlertStatus.READY_SELL1, f"+3% 매도선까지 {abs(dist_sell1):.1f}%"
        else:
            return AlertStatus.WAITING, f"대기 중"
    
    return AlertStatus.WATCHING, "관찰 중"


# ==================== 엑셀 저장 ====================
def apply_signal_formatting(file_path: str, sheet_name: str):
    """엑셀 포맷팅 적용"""
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font, Border, Side
    from openpyxl.utils import get_column_letter
    
    wb = load_workbook(file_path)
    ws = wb[sheet_name]
    
    # 테두리 스타일
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 헤더 찾기
    headers = [cell.value for cell in ws[1]]
    
    # 열 인덱스 찾기
    col_indices = {}
    for idx, header in enumerate(headers, start=1):
        col_indices[header] = idx
    
    # 금액 관련 열
    price_cols = ["종가", "저가", "고가", "20일선(당일)", "20일선(익일)", "-20%엔벨로프(당일)", "-20%엔벨로프(익일)", 
                  "1차매수선(익일)", "1차매수가(당일)", "2차매수선(익일)", "2차매수가(당일)", 
                  "3차매수선(익일)", "3차매수가(당일)", "평균매수가", 
                  "1차매도선(+3%)", "2차매도선(+5%)", "3차매도선(+7%)", "최고도달선"]
    
    # 이격도 열 (%)
    pct_cols = ["1차매수선이격도(%)", "2차매수선이격도(%)", "3차매수선이격도(%)",
                "1차매도선이격도(%)", "2차매도선이격도(%)", "3차매도선이격도(%)", "실현수익률(%)"]
    
    # 날짜 열
    date_cols = ["1차매수일", "2차매수일", "3차매수일", "종료일"]
    
    # 데이터 행 포맷팅
    for row_idx in range(2, ws.max_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            header = headers[col_idx - 1]
            
            # 테두리 적용
            cell.border = thin_border
            
            # 금액 포맷 (천 자리 콤마)
            if header in price_cols:
                if cell.value is not None and cell.value != "":
                    cell.number_format = '#,##0'
                    cell.alignment = Alignment(horizontal="right", vertical="center")
            
            # 이격도 포맷 (% 기호 포함)
            elif header in pct_cols:
                if cell.value is not None and cell.value != "":
                    # 이미 % 값이므로 % 기호만 추가
                    cell.number_format = '0.00"%"'
                    cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # 날짜 포맷
            elif header in date_cols:
                if cell.value is not None and cell.value != "":
                    cell.number_format = 'YYYY-MM-DD'
                    cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # 텍스트 중앙 정렬 (티커, 종목명, 상태 등)
            elif header in ["티커", "종목명", "매수상태", "알람상태", "종료사유"]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # 기타는 왼쪽 정렬
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
    
    # 헤더 포맷팅
    for col_idx in range(1, ws.max_column + 1):
        header_cell = ws.cell(row=1, column=col_idx)
        header = headers[col_idx - 1]
        
        # 테두리 적용
        header_cell.border = thin_border
        
        header_cell.font = Font(bold=True)
        header_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # 열 너비 자동 조정
    for col_idx in range(1, ws.max_column + 1):
        column_letter = get_column_letter(col_idx)
        header = headers[col_idx - 1]
        
        # 열 너비 설정
        if header == "티커":
            ws.column_dimensions[column_letter].width = 10
        elif header == "종목명":
            ws.column_dimensions[column_letter].width = 15
        elif header == "상태메시지":
            ws.column_dimensions[column_letter].width = 30
        elif header in date_cols:
            ws.column_dimensions[column_letter].width = 12
        elif header in price_cols:
            ws.column_dimensions[column_letter].width = 12
        elif header in pct_cols:
            ws.column_dimensions[column_letter].width = 12
        elif header == "매수상태" or header == "알람상태":
            ws.column_dimensions[column_letter].width = 12
        else:
            ws.column_dimensions[column_letter].width = 12
    
    
    wb.save(file_path)


def save_signals(df_summary: pd.DataFrame, df_history: pd.DataFrame, file_path: str):
    """시그널을 엑셀에 저장 (Summary + History)"""

    # ⭐ History 백업: History가 비어있지 않으면 백업 파일 생성
    if not df_history.empty and Path(file_path).exists():
        backup_path = file_path.replace(".xlsx", "_history_backup.xlsx")
        try:
            # 기존 파일에서 History만 백업
            df_history.to_excel(backup_path, sheet_name="History", index=False)
            logger.info(f"✓ History 백업 완료: {backup_path}")
        except Exception as e:
            logger.warning(f"⚠️ History 백업 실패: {e}")

    # ⭐ 매수량, 총투자금액, 총보유수량 열 제거
    cols_to_drop = ["1차매수량", "2차매수량", "3차매수량", "총투자금액", "총보유수량"]
    df_summary = df_summary.drop(columns=[c for c in cols_to_drop if c in df_summary.columns], errors='ignore')
    df_history = df_history.drop(columns=[c for c in cols_to_drop if c in df_history.columns], errors='ignore')

    # 엑셀에 저장
    with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
        df_summary.to_excel(writer, sheet_name="Summary", index=False)
        df_history.to_excel(writer, sheet_name="History", index=False)

    # 포맷팅 적용
    apply_signal_formatting(file_path, "Summary")
    apply_signal_formatting(file_path, "History")

    logger.info(f"✓ Summary 저장 완료: {len(df_summary)}개 종목")
    if not df_history.empty:
        logger.info(f"✓ History 저장 완료: {len(df_history)}개 종목")
    else:
        logger.warning(f"⚠️ History가 비어있습니다!")


def move_to_history(df_summary: pd.DataFrame, df_history: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """SOLD 상태 종목을 Summary → History로 이동 (축적식) 및 Turnover Universe에서 삭제"""
    
    # SOLD 종목 찾기
    mask_sold = df_summary["매수상태"] == BuyStatus.SOLD
    df_sold = df_summary[mask_sold].copy()
    
    if df_sold.empty:
        return df_summary, df_history
    
    # History에 추가 (종료일, 종료사유, 실현수익률 추가) - ⭐ 축적식으로 누적
    now = datetime.now().strftime("%Y-%m-%d")
    
    for idx, row in df_sold.iterrows():
        # 종료일 추가
        row["종료일"] = now
        
        # 종료사유 판단 (최고도달선 기준)
        max_high = row.get("최고도달선")
        sell3 = row.get("3차매도선(+7%)")
        sell2 = row.get("2차매도선(+5%)")
        sell1 = row.get("1차매도선(+3%)")
        
        if max_high and sell3 and max_high >= sell3:
            row["종료사유"] = "+7% 도달 → 전량 매도"
        elif max_high and sell2 and max_high >= sell2:
            row["종료사유"] = "+5% 도달 후 매수가 재터치 → 전량 매도"
        elif max_high and sell1 and max_high >= sell1:
            row["종료사유"] = "+3% 도달 후 매수가 재터치 → 전량 매도"
        else:
            row["종료사유"] = "매도 완료"
        
        # 실현수익률 계산
        avg_price = row.get("평균매수가")
        close = row.get("종가")
        if avg_price and close:
            profit_pct = ((close - avg_price) / avg_price) * 100
            row["실현수익률(%)"] = profit_pct
        else:
            row["실현수익률(%)"] = 0
        
        # ⭐ 축적식: 기존 History에 계속 추가 (제거 안 함)
        df_history = pd.concat([df_history, row.to_frame().T], ignore_index=True)
    
    # Summary에서 제거
    df_summary = df_summary[~mask_sold].reset_index(drop=True)
    
    # Market Cap Universe에서도 매도 완료된 종목 삭제
    try:
        turnover_file = "output/marketcap_universe.xlsx"
        if os.path.exists(turnover_file):
            df_turnover = pd.read_excel(turnover_file, dtype={'티커': str})
            sold_tickers = df_sold['티커'].tolist()
            
            # 매도 완료된 종목을 Market Cap Universe에서 삭제
            df_turnover_updated = df_turnover[~df_turnover['티커'].isin(sold_tickers)]
            
            # 파일 저장
            df_turnover_updated.to_excel(turnover_file, index=False)
            
            logger.info(f"✓ Market Cap Universe에서 {len(sold_tickers)}개 종목 삭제 완료")
        else:
            logger.warning("⚠ Market Cap Universe 파일을 찾을 수 없습니다.")
    except Exception as e:
        logger.error(f"❌ Market Cap Universe 업데이트 실패: {e}")
    
    logger.info(f"✓ {len(df_sold)}개 종목을 History로 이동 (총 {len(df_history)}개 기록)")
    
    return df_summary, df_history


# ==================== 메인 ====================
def main():
    parser = argparse.ArgumentParser(description="Trading Signal System")
    parser.add_argument("--appkey", required=True, help="Kiwoom API App Key")
    parser.add_argument("--secret", required=True, help="Kiwoom API Secret Key")
    parser.add_argument("--universe", default=DEFAULT_UNIVERSE_FILE, help="유니버스 파일 경로")
    parser.add_argument("--signal", default=DEFAULT_SIGNAL_FILE, help="시그널 파일 경로")
    parser.add_argument("--alert-threshold", type=float, default=DEFAULT_ALERT_THRESHOLD, help="알람 임계값 (퍼센트)")
    parser.add_argument("--force", action="store_true", help="거래일 체크 무시하고 강제 실행")
    
    args = parser.parse_args()
    
    universe_file = args.universe
    signal_file = args.signal
    alert_threshold = args.alert_threshold
    
    try:
        logger.info("=" * 80)
        logger.info("Trading Signal System 시작")
        logger.info(f"실행 시간: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        logger.info("=" * 80)
        
        # 거래일 체크 (강제 실행 옵션이 없는 경우에만)
        if not args.force:
            trading_info = get_trading_day_info()
            if not trading_info['is_trading_day']:
                logger.info(f"📅 비거래일입니다 ({trading_info['reason']})")
                logger.info("거래일이 아닌 날에는 시그널 분석을 건너뜁니다.")
                logger.info("강제 실행하려면 --force 옵션을 사용하세요.")
                logger.info("=" * 80)
                return
        else:
            logger.info("🔧 강제 실행 모드: 거래일 체크를 무시합니다.")
        
        # 1. API 토큰 획득
        try:
            token = get_api_token(args.appkey, args.secret)
        except Exception as e:
            logger.error(f"API 토큰 획득 실패: {e}")
            if TELEGRAM_AVAILABLE:
                send_error_alert(f"API 토큰 획득 실패: {e}", "Trading_Signal_System_S1")
            sys.exit(1)
        
        # 2. 유니버스 로드
        try:
            df_universe = pd.read_excel(universe_file, sheet_name=0)
            logger.info(f"✓ 추적 대상: {len(df_universe)}개 종목")
        except FileNotFoundError:
            logger.error(f"유니버스 파일 없음: {universe_file}")
            sys.exit(1)
        except Exception as e:
            logger.error(f"유니버스 로드 실패: {e}")
            sys.exit(1)
        
        # 3. 기존 시그널 로드 (있으면)
        df_summary = pd.DataFrame()
        df_history = pd.DataFrame()

        if Path(signal_file).exists():
            try:
                df_summary = pd.read_excel(signal_file, sheet_name="Summary")
                df_history = pd.read_excel(signal_file, sheet_name="History")
                logger.info(f"✓ 기존 시그널 로드: Summary {len(df_summary)}개, History {len(df_history)}개")

                # ⭐ History 보호: 빈 DataFrame인데 파일에 데이터가 있었다면 경고
                if df_history.empty:
                    logger.warning("⚠️ History 시트가 비어있습니다. 과거 데이터가 손실되었을 수 있습니다.")

            except Exception as e:
                logger.warning(f"기존 시그널 로드 실패 (새로 생성): {e}")
        
        # 4. 종목별 분석
        logger.info("\n" + "=" * 80)
        logger.info("종목별 분석 시작")
        logger.info("=" * 80)

        results = []
        alerts = []
        analyzed_tickers = set()

        for idx, row in df_universe.iterrows():
            ticker = str(row["티커"]).zfill(6)
            name = row["종목명"]
            recent_leading_date = row.get("최근주도주")  # ⭐ 최근 주도주 날짜 (재등장 시점)

            logger.info(f"\n[{idx + 1}/{len(df_universe)}] {name} ({ticker}) 분석 중...")

            # History에 이미 있는 종목이 재등장한 경우 (History는 유지, Summary만 신규 추가)
            if not df_history.empty and ticker in df_history["티커"].values:
                logger.info(f"  ♻️ 과거 매매 기록 있음 - 새 사이클 시작")

            result = analyze_stock(token, ticker, name, df_summary, alert_threshold)


            if result:
                results.append(result)
                analyzed_tickers.add(ticker)
        
                # 알람 대상 확인
                alert_status = result["알람상태"]
                if alert_status not in [AlertStatus.WATCHING, AlertStatus.WAITING]:
                    alerts.append(result)
                    logger.info(f"  🔔 {result['상태메시지']}")

            time.sleep(0.2)  # API 레이트 리미트
    
        # 5. Summary 업데이트 (현재 분석 결과만 저장)
        if results:
            df_summary = pd.DataFrame(results)

        # 6. SOLD 종목 History로 이동
        df_summary, df_history = move_to_history(df_summary, df_history)

        # 7. 저장
        save_signals(df_summary, df_history, signal_file)

        # 8. 알람 출력
        logger.info("\n" + "=" * 80)
        logger.info(f"🔔 알람: {len(alerts)}개")
        logger.info("=" * 80)

        for alert in alerts:
            logger.info(f"🔴 {alert['종목명']} ({alert['티커']}): {alert['상태메시지']}")

        # 9. 텔레그램 일일 리포트 전송
        if TELEGRAM_AVAILABLE:
            try:
                # 모든 사람에게 전송
                send_daily_report(alerts, len(df_summary), recipients=["all"], system_label="S1")
                logger.info("✓ 텔레그램 일일 리포트 전송 완료")
            except Exception as e:
                logger.error(f"텔레그램 전송 실패: {e}")

        # 10. 완료
        logger.info("\n" + "=" * 80)
        logger.info("완료")
        logger.info(f"분석: {len(results)}개 종목")
        logger.info(f"알람: {len(alerts)}개")
        logger.info("=" * 80)
    
    except Exception as e:
        logger.error(f"예기치 않은 오류 발생: {e}", exc_info=True)
        if TELEGRAM_AVAILABLE:
            send_error_alert(f"예기치 않은 오류: {str(e)}", "Trading_Signal_System_S1")
        sys.exit(1)


if __name__ == "__main__":
    main()

