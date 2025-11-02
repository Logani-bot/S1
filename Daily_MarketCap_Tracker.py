#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Daily Market Cap Tracker - 매일 시가총액 1조 3천억+ 종목 추적 (S1 시스템)
- 매일 1회 실행하여 오늘의 시가총액 순위 수집
- 1조 3천억 이상 종목만 엑셀에 축적
- (날짜, 티커) 중복 자동 제거
"""

import os
import sys
import argparse
import logging
import time
from datetime import datetime, date
from typing import List, Dict, Tuple, Optional

import requests
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side

# 거래일 체크 유틸리티 import
try:
    from trading_day_utils import is_trading_day, get_trading_day_info, get_previous_trading_day
    TRADING_DAY_CHECK_AVAILABLE = True
except ImportError:
    TRADING_DAY_CHECK_AVAILABLE = False
    logger_import = logging.getLogger(__name__)
    logger_import.warning("trading_day_utils 모듈을 찾을 수 없습니다. 거래일 체크가 비활성화됩니다.")

# ==================== 설정 ====================
APPKEY_DEFAULT = "IweTdkYa8JWDUOa8NohVSVeOiJ1THDGd_2x050A8XcU"
SECRET_DEFAULT = "eazu-jPNJpAsIVkaUTh3_88gUvXrCMJCwGF2AYRtBJs"

API_BASE_URL = "https://api.kiwoom.com"
API_TOKEN_URL = "https://api.kiwoom.com/oauth2/token"
API_STOCK_INFO_ENDPOINT = "/api/dostk/stkinfo"
API_STOCK_INFO_ID = "ka10099"  # 시가총액 순위용: 종목 정보 조회 API

EXCEL_PATH = "output/marketcap_universe.xlsx"
SHEET_NAME = "universe"

THRESHOLD_MARKET_CAP = 13000.0  # 1조 3천억원 (단위: 억원)
MARKETS = ["0", "10"]  # 0: 코스피, 10: 코스닥

# ETF/ETN 제외 키워드
EXCLUDE_KEYWORDS = [
    "KODEX", "TIGER", "KBSTAR", "KOSEF", "ARIRANG", "HANARO", "SOL", 
    "TREX", "ACE", "인버스", "레버리지", "선물", "ETF", "ETN", "지수"
]

# ==================== 로깅 ====================
logging.basicConfig(
    level=logging.INFO,
    format="[%(asctime)s] %(levelname)s: %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S"
)
logger = logging.getLogger(__name__)

# ==================== API 함수 ====================
def get_access_token(appkey: str, secret: str) -> str:
    """키움 API 토큰 획득"""
    headers = {"Content-Type": "application/json;charset=UTF-8"}
    body = {
        "grant_type": "client_credentials",
        "appkey": appkey.strip('"'),
        "secretkey": secret.strip('"')
    }
    
    response = requests.post(API_TOKEN_URL, headers=headers, json=body, timeout=20)
    response.raise_for_status()
    
    data = response.json()
    token = data.get("token") or data.get("access_token")
    
    if not token:
        raise RuntimeError("토큰 획득 실패")
    
    logger.info("✓ API 토큰 획득 완료")
    return token


def fetch_stock_list(token: str, market: str, max_retry: int = 5) -> dict:
    """종목 리스트 조회 (ka10099 API - 시가총액 계산용)"""
    headers = {
        "authorization": f"Bearer {token}",
        "Content-Type": "application/json;charset=UTF-8",
        "api-id": API_STOCK_INFO_ID,
        "cont-yn": "N",
        "next-key": ""
    }
    
    body = {
        "mrkt_tp": market  # 0: 코스피, 10: 코스닥
    }
    
    url = API_BASE_URL + API_STOCK_INFO_ENDPOINT
    
    for attempt in range(max_retry):
        try:
            response = requests.post(url, headers=headers, json=body, timeout=20)
            
            # Rate limit 처리
            if response.status_code == 429:
                retry_after = response.headers.get("Retry-After", 1)
                sleep_time = float(retry_after) if str(retry_after).isdigit() else (0.5 * (2 ** attempt))
                logger.warning(f"Rate limit - {sleep_time:.1f}초 대기 중...")
                time.sleep(sleep_time)
                continue
            
            # 서버 오류 재시도
            if 500 <= response.status_code < 600:
                logger.warning(f"서버 오류 {response.status_code} - 재시도 {attempt + 1}/{max_retry}")
                time.sleep(0.5 * (2 ** attempt))
                continue
            
            response.raise_for_status()
            result = response.json()
            
            # 디버깅: 응답 구조 확인 (첫 시도만)
            if attempt == 0:
                logger.debug(f"API 응답 키: {list(result.keys())}")
                
                # list 키 확인
                if "list" in result:
                    list_data = result.get("list", [])
                    logger.debug(f"list 타입: {type(list_data)}, 길이: {len(list_data) if isinstance(list_data, list) else 'N/A'}")
                    if isinstance(list_data, list) and len(list_data) > 0:
                        logger.debug(f"첫 번째 항목 키: {list(list_data[0].keys()) if isinstance(list_data[0], dict) else 'N/A'}")
                
                # return_code 확인
                if "return_code" in result:
                    return_code = result.get("return_code")
                    return_msg = result.get("return_msg", "")
                    if return_code != 0:
                        logger.warning(f"API 오류: return_code={return_code}, return_msg={return_msg}")
                    else:
                        logger.debug(f"API 성공: return_code={return_code}, return_msg={return_msg}")
            
            return result
            
        except requests.RequestException as e:
            if attempt == max_retry - 1:
                raise
            logger.warning(f"요청 실패 - 재시도 {attempt + 1}/{max_retry}: {e}")
            time.sleep(0.5 * (2 ** attempt))
    
    raise RuntimeError("최대 재시도 횟수 초과")


# ==================== 데이터 처리 ====================
def normalize_ticker(ticker: str) -> str:
    """티커 정규화 (6자리, 알파벳 포함 가능)"""
    if not ticker:
        return ""
    
    # _AL 같은 suffix 제거
    ticker = str(ticker).split("_")[0].strip()
    
    # 알파벳이 포함된 경우 (예: 0008Z0) 그대로 반환
    if any(c.isalpha() for c in ticker):
        return ticker.zfill(6)[:6]
    
    # 숫자만 있는 경우 숫자만 추출하여 6자리로 패딩
    digits = "".join(c for c in ticker if c.isdigit())
    return digits.zfill(6)[:6] if digits else ""


def is_excluded(name: str) -> bool:
    """ETF/ETN 등 제외 대상 여부"""
    if not name:
        return True
    
    name_upper = str(name).upper()
    return any(keyword.upper() in name_upper for keyword in EXCLUDE_KEYWORDS)


def parse_stock_list_response(response: dict) -> List[Dict]:
    """ka10099 API 응답에서 종목 정보 추출 및 시가총액 계산"""
    # list 키 확인
    if "list" not in response or not isinstance(response["list"], list):
        logger.warning(f"응답에서 list 데이터를 찾을 수 없습니다. 응답 키: {list(response.keys())}")
        return []
    
    data_list = response["list"]
    
    if not data_list:
        return []
    
    results = []
    
    for item in data_list:
        # 티커 추출 (code 필드)
        ticker = None
        if "code" in item and item["code"]:
            ticker = normalize_ticker(item["code"])
        elif "stk_cd" in item and item["stk_cd"]:
            ticker = normalize_ticker(item["stk_cd"])
        
        # 종목명 추출 (name 필드)
        name = None
        if "name" in item and item["name"]:
            name = str(item["name"]).strip()
        elif "stk_nm" in item and item["stk_nm"]:
            name = str(item["stk_nm"]).strip()
        
        # 상장주식수 추출 (listCount 필드)
        list_count = None
        if "listCount" in item and item["listCount"]:
            try:
                list_count = int(str(item["listCount"]).replace(",", ""))
            except (ValueError, TypeError):
                pass
        
        # 전일종가 추출 (lastPrice 필드)
        last_price = None
        if "lastPrice" in item and item["lastPrice"]:
            try:
                last_price = int(str(item["lastPrice"]).replace(",", ""))
            except (ValueError, TypeError):
                pass
        
        # 시가총액 계산: 상장주식수 × 전일종가
        market_cap = None
        if list_count and last_price:
            try:
                market_cap = int(list_count * last_price)  # 원 단위
            except (ValueError, TypeError):
                pass
        
        # 유효한 데이터만 추가 (티커, 종목명, 시가총액 모두 필요)
        if ticker and name and market_cap:
            results.append({
                "ticker": ticker,
                "name": name,
                "market_cap_won": market_cap  # 원 단위
            })
    
    return results


def filter_stocks(data: List[Dict], threshold_market_cap: float) -> pd.DataFrame:
    """시가총액 필터링 및 정렬"""
    if not data:
        return pd.DataFrame(columns=["ticker", "name", "market_cap_eok"])
    
    df = pd.DataFrame(data)
    
    # 티커 기준 중복 제거 (시가총액 최대값 사용)
    df = df.groupby(["ticker", "name"], as_index=False)["market_cap_won"].max()
    
    # 단위 변환: 원 → 억원 (÷100000000)
    df["market_cap_eok"] = df["market_cap_won"] / 100000000
    
    # ETF/ETN 제외
    df = df[~df["name"].apply(is_excluded)].copy()
    
    # 시가총액 임계값 필터 (1조 3천억 = 13000억)
    df = df[df["market_cap_eok"] >= threshold_market_cap].copy()
    
    # 시가총액 내림차순 정렬
    df = df.sort_values("market_cap_eok", ascending=False).reset_index(drop=True)
    
    return df[["ticker", "name", "market_cap_eok"]]


# ==================== 엑셀 처리 ====================
def ensure_excel_exists(path: str):
    """엑셀 파일이 없으면 생성"""
    if not os.path.exists(path):
        wb = Workbook()
        ws = wb.active
        ws.title = SHEET_NAME
        ws.append(["첫주도주", "최근주도주", "티커", "종목명", "시가총액(억)", "누적횟수"])
        wb.save(path)
        logger.info(f"✓ 새 엑셀 파일 생성: {path}")


def read_existing_data(path: str) -> pd.DataFrame:
    """기존 엑셀 데이터 읽기"""
    ensure_excel_exists(path)
    
    try:
        # 모든 컬럼 읽기 (나중에 필요한 것만 선택)
        df = pd.read_excel(path, sheet_name=SHEET_NAME, dtype={"티커": str})
    except Exception as e:
        logger.warning(f"엑셀 읽기 실패: {e}")
        return pd.DataFrame(columns=["첫주도주", "최근주도주", "티커", "종목명", "시가총액(억)", "누적횟수"])
    
    if df.empty:
        return df
    
    # Unnamed 컬럼들과 업데이트 관련 컬럼 제거
    cols_to_drop = [col for col in df.columns if "Unnamed" in str(col) or "업데이트" in str(col) or "최종" in str(col)]
    if cols_to_drop:
        df = df.drop(columns=cols_to_drop)
    
    # 구버전 호환성: "날짜" → "첫주도주", "최근주도주"
    if "날짜" in df.columns and "첫주도주" not in df.columns:
        df["첫주도주"] = df["날짜"]
        df["최근주도주"] = df["날짜"]
        df = df.drop(columns=["날짜"])
    
    # 누적횟수 열이 없으면 추가
    if "누적횟수" not in df.columns:
        df["누적횟수"] = 1
    
    # 데이터 정규화
    df["티커"] = df["티커"].apply(normalize_ticker)
    df["첫주도주"] = pd.to_datetime(df["첫주도주"], errors="coerce").dt.date
    df["최근주도주"] = pd.to_datetime(df["최근주도주"], errors="coerce").dt.date
    df["누적횟수"] = df["누적횟수"].fillna(1).astype(int)
    df = df.dropna(subset=["첫주도주"])
    
    # ⭐ 컬럼 순서 정리 (신버전 표준 순서로)
    expected_cols = ["첫주도주", "최근주도주", "티커", "종목명", "시가총액(억)", "누적횟수"]
    # 구버전 호환성: 거래대금 → 시가총액 변경
    if "거래대금(억)" in df.columns and "시가총액(억)" not in df.columns:
        df["시가총액(억)"] = df["거래대금(억)"]
        df = df.drop(columns=["거래대금(억)"])
    df = df[expected_cols]
    
    return df


def get_last_update_date(path: str) -> Optional[date]:
    """H1 셀에서 마지막 업데이트 날짜 읽기"""
    try:
        wb = load_workbook(path)
        ws = wb[SHEET_NAME]
        h1_value = ws["H1"].value
        
        if h1_value and isinstance(h1_value, str):
            # "최종 업데이트: 2025-10-12" 형태에서 날짜 추출
            if ":" in h1_value:
                # "최종 업데이트: " 제거 후 날짜 파싱
                parts = h1_value.split(":", 1)  # 첫 번째 :로만 분리
                if len(parts) > 1:
                    date_str = parts[1].strip()  # "2025-10-12"
                    return pd.to_datetime(date_str).date()
        
        return None
    except Exception as e:
        return None


def save_to_excel(path: str, df: pd.DataFrame, update_date: str = None):
    """데이터프레임을 엑셀에 저장"""
    try:
        # 기존 파일 있으면 덮어쓰기
        with pd.ExcelWriter(path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            df.to_excel(writer, index=False, sheet_name=SHEET_NAME)
    except Exception:
        # 파일이 없거나 열 수 없으면 새로 생성
        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name=SHEET_NAME)
    
    # 서식 적용 (G1 셀 포함)
    apply_formatting(path, SHEET_NAME, update_date)


def apply_formatting(path: str, sheet_name: str, update_date: str = None):
    """엑셀 서식 적용 (열 너비, 테두리, 정렬, 숫자 포맷)"""
    try:
        from openpyxl.styles import Alignment, Font
        
        wb = load_workbook(path)
        ws = wb[sheet_name]
        
        # G열 이후 (Unnamed 등) 모두 삭제
        if ws.max_column > 6:
            ws.delete_cols(7, ws.max_column - 6)
        
        # 열 너비 수동 지정
        ws.column_dimensions['A'].width = 12  # 첫주도주
        ws.column_dimensions['B'].width = 12  # 최근주도주
        ws.column_dimensions['C'].width = 8   # 티커
        ws.column_dimensions['D'].width = 16  # 종목명
        ws.column_dimensions['E'].width = 14  # 시가총액(억)
        ws.column_dimensions['F'].width = 10  # 누적횟수
        
        # 테두리
        thin_border = Border(
            top=Side(border_style="thin", color="000000"),
            left=Side(border_style="thin", color="000000"),
            right=Side(border_style="thin", color="000000"),
            bottom=Side(border_style="thin", color="000000")
        )
        
        # 데이터 행 서식
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=6), start=2):
            for col_idx, cell in enumerate(row, start=1):
                if cell.value not in (None, ""):
                    cell.border = thin_border
                    
                    # A,B열(날짜): 중앙정렬
                    if col_idx in [1, 2]:
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    
                    # C열(티커): 텍스트 포맷 + 중앙정렬 (⭐ 앞의 0 보존)
                    elif col_idx == 3:
                        cell.number_format = '@'  # 텍스트 포맷
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    
                    # D열(종목명): 중앙정렬
                    elif col_idx == 4:
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    
                    # E열(시가총액): 천단위 콤마 + 오른쪽 정렬
                    elif col_idx == 5:
                        cell.number_format = '#,##0.00'
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                    
                    # F열(누적횟수): 오른쪽 정렬
                    elif col_idx == 6:
                        cell.alignment = Alignment(horizontal="right", vertical="center")
        
        # 헤더 행 중앙정렬
        for cell in ws[1]:
            if cell.value:
                cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # H1 셀에 업데이트 날짜 기록 (pandas 컬럼 인식 방지)
        if update_date:
            ws["H1"] = f"최종 업데이트: {update_date}"
            ws["H1"].font = Font(bold=True, size=10)
            ws["H1"].alignment = Alignment(horizontal="left", vertical="center")
            ws["H1"].border = Border()  # 테두리 없음
        
        wb.save(path)
    except Exception as e:
        logger.warning(f"서식 적용 실패: {e}")


def append_to_excel(path: str, new_rows: List[Tuple[date, str, str, float]]):
    """새 데이터를 엑셀에 추가 (종목별 1행 유지 + 누적횟수 카운트)"""
    if not new_rows:
        logger.info("저장할 데이터가 없습니다.")
        return
    
    # 오늘 날짜 (스크립트 실행 날짜)
    today = date.today()
    
    # H1 셀에서 마지막 업데이트 날짜 읽기
    last_update = get_last_update_date(path)
    
    # 기존 데이터 읽기
    df_old = read_existing_data(path)
    
    # 새 데이터 생성 (API에서 받은 거래일)
    df_new = pd.DataFrame(new_rows, columns=["거래일", "티커", "종목명", "시가총액(억)"])
    df_new["티커"] = df_new["티커"].apply(normalize_ticker)
    trading_date = pd.to_datetime(df_new["거래일"]).dt.date.iloc[0]  # API가 준 거래일
    
    # 신규 종목용: 첫주도주 = 최근주도주 = 거래일
    df_new["첫주도주"] = trading_date
    df_new["최근주도주"] = trading_date
    df_new["누적횟수"] = 1
    df_new = df_new.drop(columns=["거래일"])  # 임시 컬럼 제거
    
    # 기존 종목과 신규 종목 분리
    if not df_old.empty:
        existing_tickers = set(df_old["티커"].values)
        
        # 기존에 있는 종목: 업데이트
        mask_existing = df_new["티커"].isin(existing_tickers)
        df_existing_update = df_new[mask_existing].copy()
        
        # 기존에 없는 종목: 그대로 추가
        df_really_new = df_new[~mask_existing].copy()
        
        # 기존 데이터 업데이트
        same_day_count = 0  # 같은 날 재실행 카운트
        
        for _, row in df_existing_update.iterrows():
            ticker = row["티커"]
            mask = df_old["티커"] == ticker
            
            # ⭐ 핵심: H1 셀의 업데이트날짜 기준으로 판단 (거래일이 아님!)
            # 오늘 처음 실행이면 (H1의 날짜가 오늘이 아니면) 누적횟수 +1
            if last_update != today:
                df_old.loc[mask, "누적횟수"] += 1
            else:
                # 오늘 이미 실행했었음 (같은 날 재실행)
                same_day_count += 1
            
            # 첫주도주: 유지 (변경 안 함)
            # 최근주도주: 항상 최신 거래일로 갱신
            df_old.loc[mask, "최근주도주"] = row["최근주도주"]
            df_old.loc[mask, "시가총액(억)"] = row["시가총액(억)"]
        
        # 기존 데이터 + 신규 종목만 합치기
        df_all = pd.concat([df_old, df_really_new], ignore_index=True)
        
        new_count = len(df_really_new)
        updated_count = len(df_existing_update)
    else:
        # 기존 데이터가 없으면 모두 신규
        df_all = df_new
        new_count = len(df_new)
        updated_count = 0
        same_day_count = 0
    
    # 누적횟수 내림차순 → 최근주도주 내림차순 정렬
    df_all = df_all.sort_values(["누적횟수", "최근주도주"], ascending=[False, False]).reset_index(drop=True)
    
    # H1 셀에 기록할 업데이트 날짜
    update_date = today.strftime("%Y-%m-%d")
    
    # 저장 (H1 셀 포함)
    save_to_excel(path, df_all, update_date)
    
    # 결과 로그
    if same_day_count > 0 and same_day_count == updated_count:
        logger.info(f"✓ 데이터 저장 완료: {updated_count}개 종목 갱신 (오늘 이미 실행함, 누적횟수 유지)")
    elif new_count > 0 and updated_count > 0:
        actual_increase = updated_count - same_day_count
        if actual_increase > 0:
            logger.info(f"✓ 데이터 저장 완료: 신규 {new_count}개, 누적횟수 증가 {actual_increase}개" + 
                       (f", 재실행 {same_day_count}개" if same_day_count > 0 else ""))
        else:
            logger.info(f"✓ 데이터 저장 완료: 신규 {new_count}개")
    elif new_count > 0:
        logger.info(f"✓ 데이터 저장 완료: 신규 {new_count}개 종목 추가")
    elif updated_count > 0:
        actual_increase = updated_count - same_day_count
        if actual_increase > 0:
            logger.info(f"✓ 데이터 저장 완료: {actual_increase}개 종목의 누적횟수 증가" +
                       (f" (재실행 {same_day_count}개 제외)" if same_day_count > 0 else ""))
        else:
            logger.info(f"✓ 데이터 저장 완료: {updated_count}개 종목 갱신 (오늘 이미 실행함)")
    else:
        logger.info("✓ 데이터 저장 완료: 변경 없음")


# ==================== 메인 로직 ====================
def collect_today_data(token: str, threshold_market_cap: float, excel_path: str, force_date: date = None) -> int:
    """오늘의 시가총액 데이터 수집 및 저장"""
    today = date.today()
    
    # 조회할 날짜 결정
    if force_date:
        query_date = force_date
    elif TRADING_DAY_CHECK_AVAILABLE:
        # 오늘이 거래일이 아니면 최근 거래일로 조회
        if is_trading_day(today):
            query_date = today
        else:
            query_date = get_previous_trading_day(today)
            logger.info(f"⚠️ 오늘({today})은 거래일이 아닙니다. 최근 거래일({query_date}) 데이터를 조회합니다.")
    else:
        query_date = today
    
    logger.info(f"{'='*60}")
    logger.info(f"조회 날짜: {query_date}")
    logger.info(f"임계값: {threshold_market_cap:,.0f}억 이상 (시가총액)")
    logger.info(f"{'='*60}")
    
    # 모든 마켓 데이터 수집
    all_data = []
    
    for market in MARKETS:
        market_name = "코스피" if market == "0" else "코스닥"
        logger.info(f"[{market_name}] 종목 리스트 조회 중...")
        
        try:
            response = fetch_stock_list(token, market)
            parsed = parse_stock_list_response(response)
            all_data.extend(parsed)
            logger.info(f"  ✓ {len(parsed)}개 종목 시가총액 계산 완료")
        except Exception as e:
            logger.error(f"  ✗ 조회 실패: {e}")
    
    if not all_data:
        logger.warning("조회된 데이터가 없습니다.")
        return 0
    
    # 필터링
    df_filtered = filter_stocks(all_data, threshold_market_cap)
    
    if df_filtered.empty:
        logger.warning(f"{threshold_market_cap:,.0f}억 이상 종목이 없습니다.")
        return 0
    
    # 결과 출력
    logger.info(f"\n{'='*60}")
    logger.info(f"시가총액 {threshold_market_cap:,.0f}억 이상: {len(df_filtered)}개 종목")
    logger.info(f"{'='*60}")
    
    logger.info("\n상위 10개:")
    for idx, row in df_filtered.head(10).iterrows():
        logger.info(f"  {idx+1:2d}. {row['ticker']} {row['name']:20s} {row['market_cap_eok']:>10,.0f}억")
    
    # 엑셀에 저장
    rows_to_save = [
        (today, row["ticker"], row["name"], row["market_cap_eok"])
        for _, row in df_filtered.iterrows()
    ]
    
    append_to_excel(excel_path, rows_to_save)
    
    return len(rows_to_save)


# ==================== 엔트리 포인트 ====================
def main():
    parser = argparse.ArgumentParser(
        description="매일 시가총액 1조 3천억+ 종목 추적 스크립트 (S1 시스템)"
    )
    parser.add_argument(
        "--threshold", 
        type=float, 
        default=THRESHOLD_MARKET_CAP,
        help=f"시가총액 임계값 (억원, 기본값: {THRESHOLD_MARKET_CAP:,.0f})"
    )
    parser.add_argument(
        "--out", 
        default=EXCEL_PATH,
        help=f"출력 엑셀 파일명 (기본값: {EXCEL_PATH})"
    )
    parser.add_argument(
        "--appkey",
        default=None,
        help="키움 API APPKEY (환경변수 KIWOOM_APPKEY 또는 기본값 사용)"
    )
    parser.add_argument(
        "--secret",
        default=None,
        help="키움 API SECRET (환경변수 KIWOOM_SECRET 또는 기본값 사용)"
    )
    parser.add_argument(
        "--verbose",
        action="store_true",
        help="상세 로그 출력"
    )
    parser.add_argument(
        "--force",
        action="store_true",
        help="거래일 체크 무시하고 강제 실행"
    )

    args = parser.parse_args()

    if args.verbose:
        logging.getLogger().setLevel(logging.DEBUG)

    # 거래일 체크 (강제 실행 옵션이 없는 경우에만)
    if TRADING_DAY_CHECK_AVAILABLE and not args.force:
        trading_info = get_trading_day_info()
        if not trading_info['is_trading_day']:
            logger.info("=" * 60)
            logger.info(f"📅 비거래일입니다 ({trading_info['reason']})")
            logger.info("거래일이 아닌 날에는 데이터 수집을 건너뜁니다.")
            logger.info("강제 실행하려면 --force 옵션을 사용하세요.")
            logger.info("=" * 60)
            return
    elif args.force:
        logger.info("🔧 강제 실행 모드: 거래일 체크를 무시합니다.")
    
    # API 키 설정
    appkey = args.appkey or os.getenv("KIWOOM_APPKEY") or APPKEY_DEFAULT
    secret = args.secret or os.getenv("KIWOOM_SECRET") or SECRET_DEFAULT
    
    excel_path = os.path.abspath(args.out)
    
    try:
        # 토큰 획득
        token = get_access_token(appkey, secret)
        
        # 데이터 수집 및 저장 (force 옵션이면 최근 거래일로 조회)
        force_date = None
        if args.force and TRADING_DAY_CHECK_AVAILABLE:
            force_date = get_previous_trading_day()
            logger.info(f"📅 최근 거래일({force_date}) 데이터를 조회합니다.")
        
        count = collect_today_data(token, args.threshold, excel_path, force_date=force_date)
        
        logger.info(f"\n{'='*60}")
        logger.info(f"완료: {count}개 종목 저장됨")
        logger.info(f"파일: {excel_path}")
        logger.info(f"{'='*60}")
        
    except Exception as e:
        logger.exception(f"오류 발생: {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()

